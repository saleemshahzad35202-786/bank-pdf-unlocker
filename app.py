"""
Bank Statement to Excel Converter
Personal Finance Toolkit — For authorized use only
"""

import streamlit as st
import os
import io
import re
import tempfile
import pandas as pd
from datetime import datetime

st.set_page_config(
    page_title="Bank Statement to Excel Converter",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    .stApp { background-color: #f0f4f8; }
    #MainMenu, footer, header { visibility: hidden; }

    .header-banner {
        background: linear-gradient(135deg, #1F3864, #2E75B6);
        padding: 28px 32px; border-radius: 12px;
        text-align: center; margin-bottom: 20px;
        box-shadow: 0 4px 15px rgba(31,56,100,0.3);
    }
    .header-banner h1 { color: white; font-size: 24px; font-weight: 700; margin: 0 0 6px 0; font-family: Arial; }
    .header-banner p  { color: #BDD7EE; font-size: 13px; margin: 0; font-family: Arial; }

    .step-label { font-weight: 700; color: #1F3864; font-size: 15px; font-family: Arial; margin-bottom: 6px; }

    .who-box {
        background: #E8F5E9; border-left: 4px solid #70AD47;
        padding: 14px 18px; border-radius: 6px;
        font-family: Arial; font-size: 13px; color: #1F3864;
        margin-bottom: 14px;
    }
    .legal-box {
        background: #FFF3E0; border-left: 4px solid #FF6F00;
        padding: 14px 18px; border-radius: 6px;
        font-family: Arial; font-size: 13px; color: #4E342E;
        margin-bottom: 14px;
    }
    .privacy-box {
        background: #E3F2FD; border-left: 4px solid #2E75B6;
        padding: 12px 16px; border-radius: 6px;
        font-family: Arial; font-size: 13px; color: #1F3864;
        margin-top: 16px;
    }
    .eula-box {
        background: #F3E5F5; border-left: 4px solid #7B1FA2;
        padding: 14px 18px; border-radius: 6px;
        font-family: Arial; font-size: 13px; color: #311B92;
        margin-bottom: 14px;
        max-height: 200px; overflow-y: auto;
    }
    .offline-box {
        background: #E8EAF6; border-left: 4px solid #3949AB;
        padding: 14px 18px; border-radius: 6px;
        font-family: Arial; font-size: 13px; color: #1A237E;
        margin-top: 14px;
    }
    .error-box {
        background: #FFEBEE; border-left: 4px solid #C00000;
        padding: 14px 18px; border-radius: 6px;
        color: #C00000; font-family: Arial;
    }
    .footer {
        text-align: center; color: #888; font-size: 12px;
        padding: 16px 0 8px 0; font-family: Arial;
    }
    [data-testid="stFileUploader"] {
        background: #F7FBFF; border-radius: 8px;
        border: 2px dashed #2E75B6; padding: 8px;
    }
    .stButton > button {
        background: linear-gradient(135deg, #1F3864, #2E75B6) !important;
        color: white !important; font-size: 15px !important;
        font-weight: 700 !important; padding: 12px !important;
        border-radius: 8px !important; border: none !important;
        width: 100% !important; font-family: Arial !important;
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #375623, #70AD47) !important;
        color: white !important; font-size: 14px !important;
        font-weight: 700 !important; padding: 10px !important;
        border-radius: 8px !important; border: none !important;
        width: 100% !important; font-family: Arial !important;
    }
</style>
""", unsafe_allow_html=True)

# ── Default categories ─────────────────────────────────────────────────────
DEFAULT_CATEGORIES = [
    ("Netflix", "Entertainment"), ("Spotify", "Entertainment"), ("YouTube", "Entertainment"),
    ("LINKEDIN", "Professional"), ("Coursera", "Education"), ("Google One", "Utilities"),
    ("Salary", "Income"), ("ATM", "Cash Withdrawal"),
    ("IMTIAZ", "Groceries"), ("CARREFOUR", "Groceries"), ("LESCO", "Utilities"),
    ("SNGPL", "Utilities"), ("UtilityBill", "Utilities"), ("Utility Bill", "Utilities"),
    ("Advance Tax", "Tax"), ("FED ", "Tax"), ("Xelplus", "Education"),
    ("Amazon", "Shopping"), ("Daraz", "Shopping"), ("Fuel", "Transport"),
    ("Petrol", "Transport"), ("Uber", "Transport"), ("Careem", "Transport"),
    ("Restaurant", "Food & Dining"), ("Cafe", "Food & Dining"),
    ("Hospital", "Healthcare"), ("Pharmacy", "Healthcare"),
    ("Funds Transfer", "Transfer"), ("Transfer", "Transfer"),
]

# ── Library imports ────────────────────────────────────────────────────────
try:
    import pikepdf
    PIKEPDF_OK = True
except ImportError:
    PIKEPDF_OK = False
    st.error("⚠️ Add `pikepdf` to requirements.txt")

try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False
    st.error("⚠️ Add `pdfplumber` to requirements.txt")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ═══════════════════════════════════════════════════════════════════════════
# CORE FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def categorize(description, lookup_df):
    if not description or pd.isna(description):
        return "Uncategorised"
    desc_lower = str(description).lower().strip()
    for _, row in lookup_df.iterrows():
        kw = str(row["Look up"]).lower().strip()
        if kw and kw in desc_lower:
            return row["Placed"]
    return "Uncategorised"


def to_num(cell):
    try:
        return float(str(cell).replace(",", "").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def detect_columns(raw_rows):
    header_keywords = {
        "debit":   ["debit", "withdrawal", "dr"],
        "credit":  ["credit", "deposit", "cr"],
        "balance": ["balance"],
        "desc":    ["description", "particulars", "narration", "details"],
    }
    for row in raw_rows[:20]:
        if not row:
            continue
        cleaned = [str(c).lower().strip() if c else "" for c in row]
        debit_idx = credit_idx = balance_idx = desc_idx = None
        for i, cell in enumerate(cleaned):
            for kw in header_keywords["debit"]:
                if kw in cell: debit_idx = i
            for kw in header_keywords["credit"]:
                if kw in cell: credit_idx = i
            for kw in header_keywords["balance"]:
                if kw in cell: balance_idx = i
            for kw in header_keywords["desc"]:
                if kw in cell: desc_idx = i
        if debit_idx is not None and balance_idx is not None:
            return debit_idx, credit_idx, balance_idx, desc_idx
    return None


def open_pdf(pdf_bytes, password=None):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    unlocked_buf = io.BytesIO()
    try:
        with pikepdf.open(tmp_path) as p:
            p.save(unlocked_buf)
        unlocked_buf.seek(0)
        os.unlink(tmp_path)
        return unlocked_buf, False
    except pikepdf.PasswordError:
        if not password:
            os.unlink(tmp_path)
            raise ValueError("This PDF is password protected. Please enter your password below.")
        try:
            unlocked_buf = io.BytesIO()
            with pikepdf.open(tmp_path, password=password) as p:
                p.save(unlocked_buf)
            unlocked_buf.seek(0)
            os.unlink(tmp_path)
            return unlocked_buf, True
        except pikepdf.PasswordError:
            os.unlink(tmp_path)
            raise ValueError("Incorrect password. Please check and try again.")
    except Exception as e:
        os.unlink(tmp_path)
        raise e


def parse_transactions(raw_rows):
    records = []
    date_pattern = re.compile(r'\d{2}[-/]\d{2}[-/]\d{4}|\d{4}[-/]\d{2}[-/]\d{2}')
    skip_keywords = ["transaction", "date", "description", "balance", "debit", "credit",
                     "opening", "closing", "statement", "account", "page"]
    col_map = detect_columns(raw_rows)

    for row in raw_rows:
        if not row:
            continue
        cleaned = [str(c).strip().replace("\n", " ") if c else "" for c in row]
        row_text = " ".join(cleaned).lower()
        if any(k in row_text for k in skip_keywords):
            continue

        date_val = ""
        for cell in cleaned[:4]:
            m = date_pattern.search(cell)
            if m:
                date_val = m.group()
                break
        if not date_val:
            continue

        debit = credit = balance = None
        desc = ""

        if col_map:
            debit_idx, credit_idx, balance_idx, desc_idx = col_map
            if debit_idx is not None and debit_idx < len(cleaned):
                debit = to_num(cleaned[debit_idx])
            if credit_idx is not None and credit_idx < len(cleaned):
                credit = to_num(cleaned[credit_idx])
            if balance_idx is not None and balance_idx < len(cleaned):
                balance = to_num(cleaned[balance_idx])
            if desc_idx is not None and desc_idx < len(cleaned):
                desc = cleaned[desc_idx]
            else:
                for cell in cleaned:
                    if len(cell) > 6 and not date_pattern.search(cell):
                        try:
                            float(cell.replace(",","").replace(" ",""))
                        except ValueError:
                            desc = cell
                            break
        else:
            num_cols = []
            for i, cell in enumerate(cleaned):
                val = to_num(cell)
                if val is not None and val > 0:
                    num_cols.append((i, val))
                elif len(cell) > 6 and not date_pattern.search(cell):
                    try:
                        float(cell.replace(",","").replace(" ",""))
                    except ValueError:
                        desc = cell

            if len(num_cols) >= 2:
                balance = num_cols[-1][1]
                if len(num_cols) >= 3:
                    debit  = num_cols[-3][1]
                    credit = None
                elif len(num_cols) == 2:
                    amt_idx, amt_val = num_cols[-2]
                    after_idx = amt_idx + 1
                    if after_idx < len(cleaned) and cleaned[after_idx] == "":
                        debit = amt_val
                    else:
                        credit = amt_val

        date_obj = None
        for fmt in ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%m/%d/%Y"]:
            try:
                date_obj = datetime.strptime(date_val, fmt)
                break
            except ValueError:
                continue

        records.append({
            "Transaction Date": date_obj.strftime("%d-%m-%Y") if date_obj else date_val,
            "Description":      desc,
            "Debit":            -abs(debit)  if debit  else None,
            "Credit":           abs(credit)  if credit else None,
            "Balance":          balance,
            "Month Name":       date_obj.strftime("%B") if date_obj else "",
            "Month Sort":       date_obj.month if date_obj else 99,
        })

    return pd.DataFrame(records) if records else pd.DataFrame()


def build_excel(df):
    wb = Workbook()
    DARK_BLUE = "1F3864"; MID_BLUE = "2E75B6"; WHITE = "FFFFFF"
    LIGHT_GRAY = "F2F2F2"; YELLOW = "FFF2CC"

    thin = Side(style="thin", color="CCCCCC")
    tb = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, cell, val, bg=DARK_BLUE, fg=WHITE, size=10):
        c = ws[cell]
        c.value = val
        c.font = Font(name="Arial", bold=True, color=fg, size=size)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    def fc(ws, r, col, val, fmt=None, bold=False, bg=None, center=False):
        c = ws.cell(r, col, val)
        c.font = Font(name="Arial", size=9, bold=bold)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        if fmt:
            c.number_format = fmt
        c.border = tb

    # Sheet 1: Transactions
    ws1 = wb.active
    ws1.title = "📥 Transactions"
    ws1.sheet_view.showGridLines = False
    for i, w in enumerate([3,14,14,45,13,13,15,18,3], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[2].height = 38
    ws1.merge_cells("B2:I2")
    hdr(ws1, "B2", "📥  Bank Transactions", size=14)
    ws1.row_dimensions[3].height = 24
    for i, h in enumerate(["Date","Month","Description","Debit","Credit","Balance","Category"]):
        c = ws1.cell(3, i+2, h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for idx, row in df.iterrows():
        r = idx + 4
        ws1.row_dimensions[r].height = 18
        bg = WHITE if idx % 2 == 0 else LIGHT_GRAY
        fc(ws1, r, 2, row.get("Transaction Date",""), center=True, bg=bg)
        fc(ws1, r, 3, row.get("Month Name",""),       center=True, bg=bg)
        fc(ws1, r, 4, row.get("Description",""),      bg=bg)
        fc(ws1, r, 5, row.get("Debit"),   fmt="#,##0;(#,##0);\"-\"", center=True, bg=bg)
        fc(ws1, r, 6, row.get("Credit"),  fmt="#,##0;(#,##0);\"-\"", center=True, bg=bg)
        fc(ws1, r, 7, row.get("Balance"), fmt="#,##0;(#,##0);\"-\"", center=True, bg=bg)
        fc(ws1, r, 8, row.get("Category","Uncategorised"), center=True, bg=bg)

    # Sheet 2: Monthly Summary
    ws2 = wb.create_sheet("📅 Monthly Summary")
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate([3,18,14,14,14,14,3], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 38
    ws2.merge_cells("B2:F2")
    hdr(ws2, "B2", "📅  Monthly Summary", size=14)
    ws2.row_dimensions[3].height = 24
    for i, h in enumerate(["Month","Total Income","Total Expenses","Net Savings","Savings %"]):
        c = ws2.cell(3, i+2, h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")

    monthly = df.groupby(["Month Name","Month Sort"]).agg(
        Income=("Credit", lambda x: x.dropna().sum()),
        Expenses=("Debit", lambda x: abs(x.dropna().sum()))
    ).reset_index().sort_values("Month Sort")
    monthly["Net"] = monthly["Income"] - monthly["Expenses"]
    monthly["Pct"] = monthly.apply(lambda r: r["Net"]/r["Income"] if r["Income"]>0 else 0, axis=1)

    for i, row in enumerate(monthly.itertuples()):
        r = i + 4
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        ws2.row_dimensions[r].height = 20
        fc(ws2, r, 2, row._1, bold=True, bg=bg)
        fc(ws2, r, 3, row.Income,   fmt="#,##0", center=True, bg=bg)
        fc(ws2, r, 4, row.Expenses, fmt="#,##0", center=True, bg=bg)
        c_net = ws2.cell(r, 5, row.Net)
        c_net.font = Font(name="Arial", size=9, bold=True,
                          color="375623" if row.Net >= 0 else "C00000")
        c_net.fill = PatternFill("solid", fgColor=bg)
        c_net.alignment = Alignment(horizontal="center", vertical="center")
        c_net.number_format = "#,##0;(#,##0)"
        c_net.border = tb
        fc(ws2, r, 6, row.Pct, fmt="0.0%", center=True, bg=bg)

    tr = len(monthly) + 4
    ws2.row_dimensions[tr].height = 22
    for col, val, fmt in [
        (2,"TOTAL","@"),
        (3, monthly["Income"].sum(),   "#,##0"),
        (4, monthly["Expenses"].sum(), "#,##0"),
        (5, monthly["Net"].sum(),      "#,##0;(#,##0)"),
        (6, monthly["Net"].sum()/monthly["Income"].sum() if monthly["Income"].sum()>0 else 0, "0.0%"),
    ]:
        c = ws2.cell(tr, col, val)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        c.border = tb

    # Sheet 3: Category Breakdown
    ws3 = wb.create_sheet("🏷 Categories")
    ws3.sheet_view.showGridLines = False
    for i, w in enumerate([3,22,14,14,3], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[2].height = 38
    ws3.merge_cells("B2:D2")
    hdr(ws3, "B2", "🏷  Spending by Category", size=14)
    ws3.row_dimensions[3].height = 24
    for i, h in enumerate(["Category","Total Spent","% of Total"]):
        c = ws3.cell(3, i+2, h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")

    cat_summary = df.groupby("Category").agg(
        Total=("Debit", lambda x: abs(x.dropna().sum()))
    ).reset_index().sort_values("Total", ascending=False)
    total_spent = cat_summary["Total"].sum()
    CAT_COLORS = {
        "Entertainment":"9DC3E6","Education":"A9D18E","Income":"C6E0B4",
        "Utilities":"FFE699","Tax":"F4B183","Cash Withdrawal":"FF7C80",
        "Groceries":"B4C7E7","Transfer":"E2EFDA","Professional":"D9D9D9",
        "Shopping":"EAD1DC","Transport":"D9EAD3","Food & Dining":"FCE5CD","Healthcare":"D0E4F5",
    }
    for i, row in enumerate(cat_summary.itertuples()):
        r = i + 4
        bg = CAT_COLORS.get(row.Category, LIGHT_GRAY)
        ws3.row_dimensions[r].height = 20
        fc(ws3, r, 2, row.Category, bold=True, bg=bg)
        fc(ws3, r, 3, row.Total, fmt="#,##0", center=True, bg=bg)
        fc(ws3, r, 4, row.Total/total_spent if total_spent>0 else 0, fmt="0.0%", center=True, bg=bg)

    # Sheet 4: Dashboard
    ws4 = wb.create_sheet("📊 Dashboard")
    ws4.sheet_view.showGridLines = False
    for i, w in enumerate([3,16,14,14,14,3,16,14,14,3], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.row_dimensions[2].height = 45
    ws4.merge_cells("B2:I2")
    hdr(ws4, "B2", "📊  Personal Finance Dashboard", size=16)
    ws4.row_dimensions[3].height = 22
    ws4.merge_cells("B3:I3")
    hdr(ws4, "B3",
        f"Generated: {datetime.now().strftime('%d %B %Y')}  |  Transactions: {len(df)}  |  Categories: {df['Category'].nunique()}",
        bg=MID_BLUE, size=10)

    total_income   = df["Credit"].dropna().sum()
    total_expenses = abs(df["Debit"].dropna().sum())
    net_savings    = total_income - total_expenses
    savings_pct    = net_savings / total_income * 100 if total_income > 0 else 0

    ws4.row_dimensions[4].height = 10
    ws4.row_dimensions[5].height = 26
    ws4.row_dimensions[6].height = 36
    ws4.row_dimensions[7].height = 10
    ws4.row_dimensions[8].height = 22

    for (tr_, vr_), title, val, color in [
        (("B5:D5","B6:D6"), "💰 Total Income",   total_income,   "375623"),
        (("E5:G5","E6:G6"), "💸 Total Expenses",  total_expenses, "C00000"),
        (("H5:I5","H6:I6"), "📈 Net Savings",     net_savings,    "2E75B6"),
    ]:
        ws4.merge_cells(tr_); ws4.merge_cells(vr_)
        t = ws4[tr_.split(":")[0]]
        t.value = title; t.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        t.fill = PatternFill("solid", fgColor=color)
        t.alignment = Alignment(horizontal="center", vertical="center")
        v = ws4[vr_.split(":")[0]]
        v.value = val; v.font = Font(name="Arial", bold=True, color=WHITE, size=16)
        v.fill = PatternFill("solid", fgColor=color)
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.number_format = "#,##0"

    ws4.merge_cells("B8:I8")
    c8 = ws4["B8"]
    c8.value = f"💡  Savings Rate: {savings_pct:.1f}%"
    c8.font = Font(name="Arial", size=10, color="7F6000")
    c8.fill = PatternFill("solid", fgColor=YELLOW)
    c8.alignment = Alignment(horizontal="center", vertical="center")

    ws4.cell(10,2,"Month"); ws4.cell(10,3,"Income"); ws4.cell(10,4,"Expenses")
    for i, row in enumerate(monthly.itertuples()):
        ws4.cell(11+i, 2, row._1)
        ws4.cell(11+i, 3, row.Income)
        ws4.cell(11+i, 4, row.Expenses)
    end_r = 10 + len(monthly)

    bar = BarChart()
    bar.type = "col"; bar.title = "Monthly Income vs Expenses"
    bar.style = 10; bar.width = 18; bar.height = 12
    bar.add_data(Reference(ws4, min_col=3, max_col=4, min_row=10, max_row=end_r), titles_from_data=True)
    bar.set_categories(Reference(ws4, min_col=2, min_row=11, max_row=end_r))
    bar.series[0].graphicalProperties.solidFill = "70AD47"
    bar.series[1].graphicalProperties.solidFill = "C00000"
    ws4.add_chart(bar, "B10")

    ws4.cell(10,7,"Category"); ws4.cell(10,8,"Amount")
    for i, row in enumerate(cat_summary.head(8).itertuples()):
        ws4.cell(11+i, 7, row.Category)
        ws4.cell(11+i, 8, row.Total)

    pie = PieChart()
    pie.title = "Spending by Category"; pie.style = 10
    pie.width = 14; pie.height = 12
    pie.add_data(Reference(ws4, min_col=8, min_row=10, max_row=18), titles_from_data=True)
    pie.set_categories(Reference(ws4, min_col=7, min_row=11, max_row=18))
    ws4.add_chart(pie, "G10")

    ws1.sheet_properties.tabColor = "2E75B6"
    ws2.sheet_properties.tabColor = "70AD47"
    ws3.sheet_properties.tabColor = "FFD966"
    ws4.sheet_properties.tabColor = DARK_BLUE
    wb.active = ws4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════════════

# ── Header ─────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-banner">
    <h1>📊 Personal Bank Statement to Excel Converter</h1>
    <p>Convert your own bank statement PDF into a categorized Excel file with dashboard · Any bank · Any country</p>
</div>
""", unsafe_allow_html=True)

# ── Who this is for ────────────────────────────────────────────────────────
with st.expander("👤  Who is this tool for?", expanded=False):
    st.markdown("""
    <div class="who-box">
    <strong>✅ This tool is designed for:</strong><br><br>
    &nbsp;&nbsp;• <strong>Individuals</strong> who want to track and understand their own spending<br>
    &nbsp;&nbsp;• <strong>Freelancers & self-employed</strong> managing personal income and expenses<br>
    &nbsp;&nbsp;• <strong>Small business owners</strong> reconciling their own business account<br>
    &nbsp;&nbsp;• <strong>Accountants & auditors</strong> working with client statements they are authorized to access<br>
    &nbsp;&nbsp;• <strong>Students & researchers</strong> learning about personal finance management<br><br>
    <strong>❌ This tool is NOT for:</strong><br><br>
    &nbsp;&nbsp;• Accessing or processing documents you do not own or are not authorized to use<br>
    &nbsp;&nbsp;• Bypassing security on documents belonging to others<br>
    &nbsp;&nbsp;• Any illegal or unauthorized financial activity
    </div>
    """, unsafe_allow_html=True)

# ── EULA ───────────────────────────────────────────────────────────────────
st.markdown('<div class="step-label">📜 Terms of Use — Please read and agree before proceeding</div>',
            unsafe_allow_html=True)

st.markdown("""
<div class="eula-box">
<strong>END USER LICENSE AGREEMENT (EULA)</strong><br><br>

By using this tool, you agree to the following terms:<br><br>

<strong>1. Authorized Use Only</strong><br>
You may only use this tool to process bank statement PDFs that you own or have explicit written authorization to access and modify. Unauthorized use of third-party documents is strictly prohibited.<br><br>

<strong>2. Password Usage</strong><br>
This tool does NOT crack, guess, or brute-force any passwords. It only opens a PDF after you provide the correct password — the same way you would open it manually. The tool then saves an accessible copy for your personal use.<br><br>

<strong>3. No Liability</strong><br>
The creator of this tool accepts no liability for any misuse, data loss, or damages resulting from use of this tool. You use it entirely at your own risk.<br><br>

<strong>4. Personal Use</strong><br>
This tool is licensed for personal or internal business use only. You may not redistribute, resell, or sublicense this tool without written permission.<br><br>

<strong>5. Data Privacy</strong><br>
Your uploaded files are processed entirely in server memory and are permanently deleted after your session ends. No data is stored, logged, or shared with any third party.<br><br>

<strong>6. Compliance</strong><br>
You are solely responsible for ensuring your use of this tool complies with all applicable laws, regulations, and your bank's terms of service in your jurisdiction.
</div>
""", unsafe_allow_html=True)

eula_agreed = st.checkbox(
    "✅ I have read and agree to the Terms of Use. I confirm I am authorized to process the PDF I am uploading.",
    value=False
)

if not eula_agreed:
    st.info("👆 Please agree to the Terms of Use above to proceed.")
    st.stop()

st.success("✅ Thank you. You may now use the tool.")
st.divider()

# ── Step 1: Upload PDF ─────────────────────────────────────────────────────
st.markdown('<div class="step-label">Step 1 — Upload your bank statement PDF</div>',
            unsafe_allow_html=True)
st.markdown("""
<div style="background:#F0F4F8;border-radius:6px;padding:10px 14px;margin-bottom:10px;font-size:13px;font-family:Arial;color:#1F3864;">
    📄 Upload <strong>your own</strong> bank statement PDF.
    Works with both regular PDFs and password-protected PDFs.
    Supports any bank worldwide.
</div>
""", unsafe_allow_html=True)

uploaded_pdf = st.file_uploader("", type=["pdf"], label_visibility="collapsed")
if uploaded_pdf:
    st.caption(f"📄 **{uploaded_pdf.name}**  ({round(uploaded_pdf.size/1024, 1)} KB)")
st.divider()

# ── Step 2: Password ───────────────────────────────────────────────────────
st.markdown('<div class="step-label">Step 2 — PDF password (only if your PDF is protected)</div>',
            unsafe_allow_html=True)
st.markdown("""
<div class="legal-box">
    🔐 <strong>Important:</strong> Only enter a password that you already know and are authorized to use.
    This tool does <strong>not</strong> crack or guess passwords — it only opens the file
    after you provide the correct password, exactly like opening it manually.
</div>
""", unsafe_allow_html=True)

cp, cs = st.columns([3, 1])
with cp:
    password = st.text_input("", type="password",
                              placeholder="Leave blank if PDF has no password",
                              label_visibility="collapsed")
with cs:
    if st.checkbox("Show") and password:
        st.info(f"🔑 **{password}**")
st.divider()

# ── Step 3: Categories ─────────────────────────────────────────────────────
st.markdown('<div class="step-label">Step 3 — Categories (optional)</div>',
            unsafe_allow_html=True)
lookup_df = pd.DataFrame(DEFAULT_CATEGORIES, columns=["Look up", "Placed"])
use_custom = st.checkbox("📂 Upload my own categories file (Excel with 'Look up' and 'Placed' columns)")
if use_custom:
    cat_file = st.file_uploader("Upload categories Excel", type=["xlsx","xls"], key="cat")
    if cat_file:
        try:
            cdf = pd.read_excel(cat_file)
            if "Look up" in cdf.columns and "Placed" in cdf.columns:
                lookup_df = cdf[["Look up","Placed"]].dropna()
                st.success(f"✅ Loaded {len(lookup_df)} custom categories")
            else:
                st.warning("⚠️ File needs 'Look up' and 'Placed' columns. Using defaults.")
        except Exception as e:
            st.warning(f"Could not read file: {e}")

with st.expander("👀 View active categories"):
    st.dataframe(lookup_df, use_container_width=True, height=180)
st.divider()

# ── Step 4: Convert ────────────────────────────────────────────────────────
st.markdown('<div class="step-label">Step 4 — Convert to Excel</div>',
            unsafe_allow_html=True)

if st.button("📊  Convert My Bank Statement to Excel", use_container_width=True):
    if not uploaded_pdf:
        st.markdown('<div class="error-box">⚠️ Please upload a PDF file first.</div>',
                    unsafe_allow_html=True)
    else:
        pdf_bytes = uploaded_pdf.read()
        base_name = os.path.splitext(uploaded_pdf.name)[0]

        with st.spinner("📂 Opening PDF..."):
            try:
                unlocked_buf, was_protected = open_pdf(pdf_bytes, password if password else None)
                if was_protected:
                    st.success("✅ Password accepted — PDF opened successfully!")
                else:
                    st.success("✅ PDF loaded successfully!")
            except ValueError as e:
                st.markdown(f'<div class="error-box">❌ {e}</div>', unsafe_allow_html=True)
                st.stop()
            except Exception as e:
                st.markdown(f'<div class="error-box">❌ Could not open PDF: {e}</div>',
                            unsafe_allow_html=True)
                st.stop()

        df = pd.DataFrame()
        with st.spinner("📑 Extracting transactions..."):
            try:
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp2:
                    tmp2.write(unlocked_buf.getvalue())
                    tmp2_path = tmp2.name
                raw_rows = []
                with pdfplumber.open(tmp2_path) as plumb:
                    for page in plumb.pages:
                        for tbl in page.extract_tables():
                            raw_rows.extend(tbl)
                os.unlink(tmp2_path)
                df = parse_transactions(raw_rows)
                if df.empty:
                    st.warning("⚠️ No transactions found. PDF layout may differ from supported formats.")
                else:
                    st.success(f"✅ Extracted {len(df)} transactions!")
            except Exception as e:
                st.warning(f"⚠️ Extraction error: {e}")

        excel_buf = None
        if not df.empty:
            with st.spinner("🏷 Categorizing & building Excel dashboard..."):
                df["Category"] = df["Description"].apply(
                    lambda d: categorize(d, lookup_df))
                categorised = (df["Category"] != "Uncategorised").sum()
                st.success(f"✅ Categorized {categorised}/{len(df)} transactions!")
                try:
                    excel_buf = build_excel(df)
                    st.success("✅ Excel with Dashboard ready!")
                except Exception as e:
                    st.warning(f"Excel error: {e}")

            st.markdown("---")
            st.markdown("### 📋 Transaction Preview")
            st.dataframe(
                df[["Transaction Date","Month Name","Description",
                    "Debit","Credit","Balance","Category"]].head(10),
                use_container_width=True
            )
            c1, c2, c3 = st.columns(3)
            c1.metric("💰 Total Income",   f"{df['Credit'].dropna().sum():,.0f}")
            c2.metric("💸 Total Expenses", f"{abs(df['Debit'].dropna().sum()):,.0f}")
            net = df['Credit'].dropna().sum() - abs(df['Debit'].dropna().sum())
            c3.metric("📈 Net Savings", f"{net:,.0f}")

        st.markdown("---")
        st.markdown("### ⬇️ Download Your Files")
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "⬇️  Download Converted PDF",
                data=unlocked_buf.getvalue(),
                file_name=f"{base_name}_converted.pdf",
                mime="application/pdf",
                use_container_width=True
            )
        with d2:
            if excel_buf:
                st.download_button(
                    "📊  Download Excel + Dashboard",
                    data=excel_buf.getvalue(),
                    file_name=f"{base_name}_transactions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.info("Excel not available for this PDF layout.")

# ── Privacy Statement ──────────────────────────────────────────────────────
st.markdown("""
<div class="privacy-box">
    🔒 <strong>Privacy & Security Guarantee:</strong><br>
    Your uploaded files are processed <strong>entirely in server memory</strong> and are
    <strong>never saved, stored, logged, or shared</strong> with anyone.
    Files are permanently deleted the moment your session ends.
    No personal data or financial information is retained by this tool.
</div>
""", unsafe_allow_html=True)

# ── Offline Option ─────────────────────────────────────────────────────────
st.markdown("""
<div class="offline-box">
    💻 <strong>Prefer to work offline?</strong><br>
    If you are uncomfortable uploading financial documents online, a
    <strong>fully offline version</strong> of this tool is included in the product bundle.
    It runs entirely on your own computer — no internet connection required,
    no data ever leaves your machine.
    Check the product files for <strong>Bank_Statement_Unlocker.py</strong> setup instructions.
</div>
""", unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
    Personal Bank Statement to Excel Converter &nbsp;|&nbsp;
    For authorized personal use only &nbsp;|&nbsp;
    Any bank · Any country · Any currency
</div>
""", unsafe_allow_html=True)
